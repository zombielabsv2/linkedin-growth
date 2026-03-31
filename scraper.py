"""
LinkedIn Growth Scraper
=======================
A tool for collecting, importing, and analyzing LinkedIn post metrics.

Modes:
  --scrape     Attempt to scrape public LinkedIn posts (may be blocked by LinkedIn)
  --manual     CLI to manually enter post metrics
  --csv-import Import from LinkedIn's per-post analytics Excel exports (xlsx)
  --report     Print engagement summary from collected data

Usage:
  python scraper.py --scrape --profile https://www.linkedin.com/in/rahuljindal/
  python scraper.py --manual
  python scraper.py --csv-import C:\\Users\\rahul\\Downloads\\linkedin-exports
  python scraper.py --report
"""

import argparse
import csv
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR / "data"
CSV_PATH = DATA_DIR / "scraped_posts.csv"
JSON_PATH = DATA_DIR / "scraped_posts.json"

CSV_FIELDS = [
    "post_text",
    "date",
    "reactions",
    "comments",
    "impressions",
    "post_url",
    "source",          # scrape | manual | xlsx-import
    "collected_at",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;"
        "q=0.9,image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Cache-Control": "max-age=0",
}

DEFAULT_PROFILE = "https://www.linkedin.com/in/rahuljindal/"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def ensure_data_dir():
    """Create data directory if it doesn't exist."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def load_existing_posts() -> list[dict]:
    """Load existing posts from JSON, return empty list if none."""
    if JSON_PATH.exists():
        try:
            with open(JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return []
    return []


def save_posts(posts: list[dict]):
    """Persist posts to both CSV and JSON."""
    ensure_data_dir()

    # --- JSON ---
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(posts, f, indent=2, ensure_ascii=False, default=str)

    # --- CSV ---
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for p in posts:
            writer.writerow(p)

    print(f"\nSaved {len(posts)} posts to:")
    print(f"  CSV:  {CSV_PATH}")
    print(f"  JSON: {JSON_PATH}")


def dedup_posts(posts: list[dict]) -> list[dict]:
    """Remove duplicates based on (post_text, date) or post_url."""
    seen_urls = set()
    seen_text_date = set()
    unique = []
    for p in posts:
        url = (p.get("post_url") or "").strip()
        text_date = (p.get("post_text", "")[:100], p.get("date", ""))

        if url and url in seen_urls:
            continue
        if text_date in seen_text_date:
            continue

        if url:
            seen_urls.add(url)
        seen_text_date.add(text_date)
        unique.append(p)
    return unique


def safe_int(val) -> int:
    """Convert a value to int, returning 0 on failure."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    val_str = str(val).strip().replace(",", "").replace(" ", "")
    # Handle K/M suffixes
    multiplier = 1
    if val_str.lower().endswith("k"):
        multiplier = 1000
        val_str = val_str[:-1]
    elif val_str.lower().endswith("m"):
        multiplier = 1_000_000
        val_str = val_str[:-1]
    try:
        return int(float(val_str) * multiplier)
    except (ValueError, TypeError):
        return 0


def normalize_date(raw: str) -> str:
    """Try to parse various date formats into YYYY-MM-DD."""
    raw = raw.strip()
    if not raw:
        return ""

    # Already YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        return raw

    # Try common formats
    for fmt in (
        "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y",
        "%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y",
        "%Y/%m/%d", "%d.%m.%Y",
        "%d/%m/%y", "%m/%d/%y",
    ):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    # Relative dates: "2 days ago", "1 week ago", etc.
    rel = re.match(r"(\d+)\s*(day|week|month|year)s?\s*ago", raw, re.IGNORECASE)
    if rel:
        n = int(rel.group(1))
        unit = rel.group(2).lower()
        now = datetime.now()
        if unit == "day":
            return (now - timedelta(days=n)).strftime("%Y-%m-%d")
        elif unit == "week":
            return (now - timedelta(weeks=n)).strftime("%Y-%m-%d")
        elif unit == "month":
            return (now - timedelta(days=n * 30)).strftime("%Y-%m-%d")
        elif unit == "year":
            return (now - timedelta(days=n * 365)).strftime("%Y-%m-%d")

    return raw  # Return as-is if unparseable


# ---------------------------------------------------------------------------
# 1. Scrape Mode
# ---------------------------------------------------------------------------

def scrape_profile(profile_url: str) -> list[dict]:
    """
    Attempt to scrape publicly visible LinkedIn posts.

    LinkedIn aggressively blocks unauthenticated scraping, so this will
    likely return an auth wall. When that happens, the function returns
    an empty list and prints guidance.
    """
    import requests
    from bs4 import BeautifulSoup

    profile_url = profile_url.rstrip("/")
    posts_url = f"{profile_url}/recent-activity/all/"

    print(f"Attempting to fetch: {posts_url}")
    print("(LinkedIn blocks most unauthenticated requests — this may not return data)\n")

    posts = []

    try:
        session = requests.Session()
        session.headers.update(HEADERS)

        # Initial page load to get cookies
        time.sleep(1)
        resp = session.get(posts_url, timeout=15, allow_redirects=True)
        print(f"HTTP {resp.status_code} — {len(resp.text)} bytes")

        if resp.status_code == 429:
            print("\nRate limited by LinkedIn. Try again later.")
            return []

        if resp.status_code == 999:
            print("\nLinkedIn returned 999 (anti-scraping block).")
            print("This is expected — LinkedIn requires authentication for post data.")
            _print_fallback_guidance()
            return []

        if resp.status_code != 200:
            print(f"\nUnexpected status {resp.status_code}.")
            _print_fallback_guidance()
            return []

        soup = BeautifulSoup(resp.text, "html.parser")

        # Check for auth wall
        if soup.find("form", {"class": "login"}) or "authwall" in resp.url:
            print("\nLinkedIn redirected to login / auth wall.")
            _print_fallback_guidance()
            return []

        # Try to find post containers (LinkedIn markup changes frequently)
        # These selectors target the public profile activity page
        post_containers = (
            soup.find_all("div", {"class": re.compile(r"feed-shared-update")})
            or soup.find_all("li", {"class": re.compile(r"profile-creator-shared-feed")})
            or soup.find_all("div", {"class": re.compile(r"occludable-update")})
            or soup.find_all("article")
        )

        if not post_containers:
            print(f"\nNo post containers found in the HTML ({len(resp.text)} bytes).")
            # Try a JSON-LD approach
            scripts = soup.find_all("script", {"type": "application/ld+json"})
            for s in scripts:
                try:
                    data = json.loads(s.string)
                    if isinstance(data, dict) and data.get("@type") == "Article":
                        posts.append({
                            "post_text": (data.get("headline") or data.get("description") or "")[:200],
                            "date": normalize_date(data.get("datePublished", "")),
                            "reactions": 0,
                            "comments": 0,
                            "impressions": 0,
                            "post_url": data.get("url", ""),
                            "source": "scrape",
                            "collected_at": datetime.now().isoformat(),
                        })
                except (json.JSONDecodeError, TypeError):
                    continue

            if not posts:
                _print_fallback_guidance()
            return posts

        print(f"Found {len(post_containers)} post containers. Parsing...")

        for container in post_containers:
            text_el = (
                container.find("span", {"class": re.compile(r"break-words")})
                or container.find("div", {"class": re.compile(r"feed-shared-text")})
                or container.find("p")
            )
            post_text = text_el.get_text(strip=True)[:200] if text_el else ""

            date_el = (
                container.find("time")
                or container.find("span", {"class": re.compile(r"feed-shared-actor__sub")})
            )
            raw_date = ""
            if date_el:
                raw_date = date_el.get("datetime", "") or date_el.get_text(strip=True)
            date = normalize_date(raw_date)

            reactions_el = container.find(
                "span", {"class": re.compile(r"social-counts-reactions|reactions-count")}
            )
            reactions = safe_int(reactions_el.get_text(strip=True)) if reactions_el else 0

            comments_el = container.find(
                "button", {"class": re.compile(r"comment")}
            )
            comments = 0
            if comments_el:
                nums = re.findall(r"[\d,]+", comments_el.get_text())
                if nums:
                    comments = safe_int(nums[0])

            link_el = container.find("a", href=re.compile(r"linkedin\.com/feed/update"))
            post_url = link_el["href"] if link_el else ""

            posts.append({
                "post_text": post_text,
                "date": date,
                "reactions": reactions,
                "comments": comments,
                "impressions": 0,
                "post_url": post_url,
                "source": "scrape",
                "collected_at": datetime.now().isoformat(),
            })

            # Be polite
            time.sleep(0.5)

    except requests.exceptions.Timeout:
        print("\nRequest timed out.")
        _print_fallback_guidance()
    except requests.exceptions.ConnectionError as e:
        print(f"\nConnection error: {e}")
        _print_fallback_guidance()
    except Exception as e:
        print(f"\nUnexpected error: {e}")
        _print_fallback_guidance()

    return posts


def _print_fallback_guidance():
    """Print instructions for alternative data collection methods."""
    print("\n" + "=" * 70)
    print("LinkedIn blocked the scrape. Use one of these alternatives:")
    print("=" * 70)
    print()
    print("  1. MANUAL ENTRY (quick for a few posts):")
    print("     python scraper.py --manual")
    print()
    print("  2. XLSX IMPORT (best for bulk data):")
    print("     - Go to linkedin.com > Me > Posts & Activity")
    print("     - Or: linkedin.com/analytics/post-analytics/")
    print("     - Click 'Export' to download per-post analytics as Excel")
    print("     - Then run:")
    print("     python scraper.py --csv-import <folder-with-xlsx-files>")
    print()
    print("  3. LINKEDIN DATA EXPORT (most complete):")
    print("     - Go to Settings > Data Privacy > Get a copy of your data")
    print("     - Select 'Posts' and download the archive")
    print("     - Extract the xlsx files, then use --csv-import")
    print()


# ---------------------------------------------------------------------------
# 2. Manual Entry Mode
# ---------------------------------------------------------------------------

def manual_entry_mode():
    """Interactive CLI for manually entering post metrics."""
    print("\n" + "=" * 50)
    print("  LinkedIn Post Manual Entry")
    print("=" * 50)
    print("Enter post details one by one. Type 'done' to finish.\n")

    existing = load_existing_posts()
    new_posts = []

    while True:
        print(f"\n--- Post #{len(new_posts) + 1} ---")

        post_text = input("Post text (first line or title, or 'done'): ").strip()
        if post_text.lower() == "done":
            break

        date = input("Date (YYYY-MM-DD or any format): ").strip()
        date = normalize_date(date) if date else datetime.now().strftime("%Y-%m-%d")

        impressions = input("Impressions (0 if unknown): ").strip()
        reactions = input("Reactions / Likes: ").strip()
        comments = input("Comments: ").strip()
        post_url = input("Post URL (optional): ").strip()

        new_posts.append({
            "post_text": post_text[:200],
            "date": date,
            "reactions": safe_int(reactions),
            "comments": safe_int(comments),
            "impressions": safe_int(impressions),
            "post_url": post_url,
            "source": "manual",
            "collected_at": datetime.now().isoformat(),
        })

        print(f"  Added. ({len(new_posts)} new posts so far)")

    if not new_posts:
        print("\nNo posts entered.")
        return

    all_posts = dedup_posts(existing + new_posts)
    save_posts(all_posts)


# ---------------------------------------------------------------------------
# 3. XLSX Import Mode
# ---------------------------------------------------------------------------

def xlsx_import_mode(directory: str):
    """
    Import LinkedIn per-post analytics from Excel files.

    LinkedIn's analytics export typically has columns like:
      Date, Post Link/URL, Impressions, Reactions, Comments,
      Reposts, Engagement rate, Post content/Title, etc.

    This function tries to auto-detect the relevant columns by name matching.
    """
    import openpyxl

    dir_path = Path(directory).resolve()
    if not dir_path.exists():
        print(f"Error: Directory not found: {dir_path}")
        sys.exit(1)

    xlsx_files = list(dir_path.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No .xlsx files found in {dir_path}")
        sys.exit(1)

    print(f"\nFound {len(xlsx_files)} Excel file(s) in {dir_path}:")
    for f in xlsx_files:
        print(f"  - {f.name}")

    existing = load_existing_posts()
    imported = []

    # Column name patterns to match (case-insensitive)
    COL_PATTERNS = {
        "post_text": [
            "post content", "content", "title", "post title", "text",
            "post text", "description", "headline", "post",
        ],
        "date": [
            "date", "posted date", "post date", "published",
            "created", "created date", "published date", "publish date",
        ],
        "reactions": [
            "reactions", "likes", "reaction", "like",
            "total reactions",
        ],
        "comments": [
            "comments", "comment", "total comments",
        ],
        "impressions": [
            "impressions", "views", "impression", "view",
            "total impressions", "total views", "unique views",
        ],
        "post_url": [
            "post link", "url", "link", "post url",
            "content link", "permalink",
        ],
    }

    def match_column(header: str) -> str | None:
        """Match a header string to a known field."""
        h = header.lower().strip()
        for field, patterns in COL_PATTERNS.items():
            for pat in patterns:
                if pat == h or pat in h:
                    return field
        return None

    for xlsx_file in xlsx_files:
        print(f"\nProcessing: {xlsx_file.name}")
        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception as e:
            print(f"  Error opening file: {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row (first row with multiple non-empty cells)
            header_row_idx = None
            for i, row in enumerate(rows):
                non_empty = sum(1 for c in row if c is not None and str(c).strip())
                if non_empty >= 2:
                    header_row_idx = i
                    break

            if header_row_idx is None:
                print(f"  Sheet '{sheet_name}': no usable header row found.")
                continue

            headers = [str(c).strip() if c else "" for c in rows[header_row_idx]]
            col_map = {}  # field_name -> column_index
            for idx, h in enumerate(headers):
                if not h:
                    continue
                field = match_column(h)
                if field and field not in col_map:
                    col_map[field] = idx

            if not col_map:
                print(f"  Sheet '{sheet_name}': could not match any columns.")
                print(f"    Headers found: {headers}")
                continue

            print(f"  Sheet '{sheet_name}': matched columns: {col_map}")

            data_rows = rows[header_row_idx + 1:]
            sheet_count = 0
            for row in data_rows:
                if not row or all(c is None for c in row):
                    continue

                def get_val(field: str, default=""):
                    idx = col_map.get(field)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        return row[idx]
                    return default

                raw_date = get_val("date", "")
                if isinstance(raw_date, datetime):
                    date_str = raw_date.strftime("%Y-%m-%d")
                else:
                    date_str = normalize_date(str(raw_date))

                post_text = str(get_val("post_text", ""))[:200]
                post_url = str(get_val("post_url", ""))

                imported.append({
                    "post_text": post_text,
                    "date": date_str,
                    "reactions": safe_int(get_val("reactions", 0)),
                    "comments": safe_int(get_val("comments", 0)),
                    "impressions": safe_int(get_val("impressions", 0)),
                    "post_url": post_url,
                    "source": "xlsx-import",
                    "collected_at": datetime.now().isoformat(),
                })
                sheet_count += 1

            print(f"    Imported {sheet_count} posts from sheet '{sheet_name}'.")

        wb.close()

    if not imported:
        print("\nNo posts were imported from the Excel files.")
        print("Make sure the files contain LinkedIn analytics data with recognizable headers.")
        return

    print(f"\nTotal imported from Excel: {len(imported)} posts")
    all_posts = dedup_posts(existing + imported)
    save_posts(all_posts)


# ---------------------------------------------------------------------------
# 4. Report Mode
# ---------------------------------------------------------------------------

def report_mode():
    """Read collected data and print an engagement summary."""
    posts = load_existing_posts()
    if not posts:
        print("\nNo data found. Collect some posts first:")
        print("  python scraper.py --manual")
        print("  python scraper.py --csv-import <directory>")
        print("  python scraper.py --scrape")
        return

    print("\n" + "=" * 70)
    print(f"  LinkedIn Post Analytics Report  ({len(posts)} posts)")
    print("=" * 70)

    # --- Basic Stats ---
    total_reactions = sum(safe_int(p.get("reactions", 0)) for p in posts)
    total_comments = sum(safe_int(p.get("comments", 0)) for p in posts)
    total_impressions = sum(safe_int(p.get("impressions", 0)) for p in posts)

    print(f"\n  Total Posts:       {len(posts)}")
    print(f"  Total Reactions:   {total_reactions:,}")
    print(f"  Total Comments:    {total_comments:,}")
    if total_impressions > 0:
        print(f"  Total Impressions: {total_impressions:,}")
    print(f"  Avg Reactions:     {total_reactions / len(posts):.1f} per post")
    print(f"  Avg Comments:      {total_comments / len(posts):.1f} per post")
    if total_impressions > 0:
        print(f"  Avg Impressions:   {total_impressions / len(posts):,.0f} per post")
        print(f"  Avg Eng. Rate:     {(total_reactions + total_comments) / total_impressions * 100:.2f}%")

    # --- Top Posts by Engagement (reactions + comments) ---
    ranked = sorted(
        posts,
        key=lambda p: safe_int(p.get("reactions", 0)) + safe_int(p.get("comments", 0)),
        reverse=True,
    )

    print(f"\n{'-' * 70}")
    print("  TOP 10 POSTS BY ENGAGEMENT (reactions + comments)")
    print(f"{'-' * 70}")
    for i, p in enumerate(ranked[:10], 1):
        eng = safe_int(p.get("reactions", 0)) + safe_int(p.get("comments", 0))
        text = (p.get("post_text") or "(no text)")[:80]
        date = p.get("date", "?")
        r = safe_int(p.get("reactions", 0))
        c = safe_int(p.get("comments", 0))
        imp = safe_int(p.get("impressions", 0))
        imp_str = f"  imp:{imp:,}" if imp else ""
        print(f"\n  {i:>2}. [{date}] {r} reactions, {c} comments{imp_str}")
        print(f"      {text}")
        url = p.get("post_url", "")
        if url:
            print(f"      {url}")

    # --- Top Posts by Impressions (if available) ---
    has_impressions = any(safe_int(p.get("impressions", 0)) > 0 for p in posts)
    if has_impressions:
        ranked_imp = sorted(
            posts,
            key=lambda p: safe_int(p.get("impressions", 0)),
            reverse=True,
        )
        print(f"\n{'-' * 70}")
        print("  TOP 10 POSTS BY IMPRESSIONS")
        print(f"{'-' * 70}")
        for i, p in enumerate(ranked_imp[:10], 1):
            text = (p.get("post_text") or "(no text)")[:80]
            date = p.get("date", "?")
            imp = safe_int(p.get("impressions", 0))
            r = safe_int(p.get("reactions", 0))
            c = safe_int(p.get("comments", 0))
            eng_rate = ""
            if imp > 0:
                eng_rate = f"  eng:{(r + c) / imp * 100:.1f}%"
            print(f"\n  {i:>2}. [{date}] {imp:,} impressions, {r} reactions, {c} comments{eng_rate}")
            print(f"      {text}")

    # --- Posting Frequency ---
    dated_posts = [p for p in posts if p.get("date") and re.match(r"\d{4}-\d{2}-\d{2}", p.get("date", ""))]
    if dated_posts:
        dates = sorted(set(p["date"] for p in dated_posts))
        print(f"\n{'-' * 70}")
        print("  POSTING FREQUENCY")
        print(f"{'-' * 70}")
        print(f"  Date range: {dates[0]} to {dates[-1]}")

        try:
            first = datetime.strptime(dates[0], "%Y-%m-%d")
            last = datetime.strptime(dates[-1], "%Y-%m-%d")
            span_days = max((last - first).days, 1)
            posts_per_week = len(dated_posts) / (span_days / 7) if span_days >= 7 else len(dated_posts)
            print(f"  Span: {span_days} days")
            print(f"  Posts per week: {posts_per_week:.1f}")
        except ValueError:
            pass

        # Monthly breakdown
        monthly = {}
        for p in dated_posts:
            month_key = p["date"][:7]  # YYYY-MM
            monthly.setdefault(month_key, []).append(p)

        if len(monthly) > 1:
            print(f"\n  Monthly Breakdown:")
            print(f"  {'Month':<10} {'Posts':>6} {'Reactions':>10} {'Comments':>9} {'Avg Eng':>8}")
            print(f"  {'-' * 45}")
            for month in sorted(monthly.keys()):
                mp = monthly[month]
                mr = sum(safe_int(x.get("reactions", 0)) for x in mp)
                mc = sum(safe_int(x.get("comments", 0)) for x in mp)
                avg_eng = (mr + mc) / len(mp) if mp else 0
                print(f"  {month:<10} {len(mp):>6} {mr:>10,} {mc:>9,} {avg_eng:>8.1f}")

    # --- Engagement Trend ---
    if len(dated_posts) >= 5:
        dated_sorted = sorted(dated_posts, key=lambda p: p["date"])
        mid = len(dated_sorted) // 2
        first_half = dated_sorted[:mid]
        second_half = dated_sorted[mid:]

        avg_first = sum(safe_int(p.get("reactions", 0)) + safe_int(p.get("comments", 0)) for p in first_half) / len(first_half)
        avg_second = sum(safe_int(p.get("reactions", 0)) + safe_int(p.get("comments", 0)) for p in second_half) / len(second_half)

        print(f"\n{'-' * 70}")
        print("  ENGAGEMENT TREND")
        print(f"{'-' * 70}")
        print(f"  First half avg engagement:  {avg_first:.1f}")
        print(f"  Second half avg engagement: {avg_second:.1f}")

        if avg_first > 0:
            change = (avg_second - avg_first) / avg_first * 100
            direction = "UP" if change > 0 else "DOWN"
            print(f"  Trend: {direction} {abs(change):.1f}%")
        elif avg_second > avg_first:
            print(f"  Trend: UP (from zero base)")

    # --- Source Breakdown ---
    sources = {}
    for p in posts:
        src = p.get("source", "unknown")
        sources[src] = sources.get(src, 0) + 1

    print(f"\n{'-' * 70}")
    print("  DATA SOURCES")
    print(f"{'-' * 70}")
    for src, count in sorted(sources.items(), key=lambda x: -x[1]):
        print(f"  {src:<15} {count:>5} posts")

    print()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="LinkedIn Growth Scraper — collect and analyze post metrics",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scraper.py --scrape
  python scraper.py --scrape --profile https://www.linkedin.com/in/someone/
  python scraper.py --manual
  python scraper.py --csv-import C:\\Users\\rahul\\Downloads\\linkedin-exports
  python scraper.py --report
        """,
    )

    parser.add_argument(
        "--scrape", action="store_true",
        help="Attempt to scrape public LinkedIn posts (may be blocked)",
    )
    parser.add_argument(
        "--profile", type=str, default=DEFAULT_PROFILE,
        help=f"LinkedIn profile URL (default: {DEFAULT_PROFILE})",
    )
    parser.add_argument(
        "--manual", action="store_true",
        help="Open manual entry CLI to type in post metrics",
    )
    parser.add_argument(
        "--csv-import", type=str, metavar="DIR",
        help="Import from LinkedIn analytics xlsx files in the given directory",
    )
    parser.add_argument(
        "--report", action="store_true",
        help="Print engagement summary from collected data",
    )

    args = parser.parse_args()

    # Default to --report if no flags given and data exists
    if not any([args.scrape, args.manual, args.csv_import, args.report]):
        if JSON_PATH.exists():
            print("No mode specified. Showing report (use --help for options).\n")
            args.report = True
        else:
            parser.print_help()
            print("\n\nQuick start:")
            print("  python scraper.py --manual        # Enter post data by hand")
            print("  python scraper.py --csv-import .   # Import from Excel files")
            sys.exit(0)

    ensure_data_dir()

    # Execute modes (multiple can be combined)
    if args.scrape:
        scraped = scrape_profile(args.profile)
        if scraped:
            existing = load_existing_posts()
            all_posts = dedup_posts(existing + scraped)
            save_posts(all_posts)
        else:
            print("\nNo posts scraped. Try --manual or --csv-import instead.")

    if args.csv_import:
        xlsx_import_mode(args.csv_import)

    if args.manual:
        manual_entry_mode()

    if args.report:
        report_mode()


if __name__ == "__main__":
    main()
